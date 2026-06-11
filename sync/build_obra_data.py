#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_obra_data.py — Genera obra_data.json para el Cockpit Comuneros.
Reemplaza a sincronizarObraDesdeXlsx (Apps Script, timeouts cronicos con el
ANALISIS de 12MB). Aqui el parseo toma segundos con openpyxl.

Fuentes (Drive, via Service Account):
  - ANALISIS COMUNEROS.xlsx  -> hojas DASHBOARD, _DashData, CRONORGRAMA_AVANCE
  - (el AVANCE DIARIO ya fluye por AD_LIVIANO; no se necesita aqui)

Salida: obra_data.json con el MISMO shape que consume el frontend:
  { proyecto, ejecutadoGlobal, actividades[], curvaS{semanas,fechas,programado,real},
    dashboardKpis{}, capitulos[], _generado, fuente }
"""
import json, sys, os, io, datetime

ANALISIS_ID = "1KCl2nQ8M3h50282sLXIKqMmTlGbdJzv2"
AVANCE_ID   = "1JF2-xUey-IKiij4MFCPRhfy4LP4JqJbw"
OUT = os.path.join(os.path.dirname(__file__), "..", "obra_data.json")

def descargar_drive(file_id, dest):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    key = json.loads(os.environ["GDRIVE_SA_KEY"])
    creds = service_account.Credentials.from_service_account_info(
        key, scopes=["https://www.googleapis.com/auth/drive.readonly"])
    drive = build("drive", "v3", credentials=creds)
    meta = drive.files().get(fileId=file_id, fields="name,modifiedTime").execute()
    req = drive.files().get_media(fileId=file_id)
    buf = io.FileIO(dest, "wb")
    dl = MediaIoBaseDownload(buf, req, chunksize=10*1024*1024)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.close()
    return meta

def f2(v, d=0.0):
    try: return float(v)
    except (TypeError, ValueError): return d

def fecha_str(v):
    if isinstance(v, datetime.datetime): return v.strftime("%d/%m/%Y")
    return str(v or "")

def construir(xlsx_path, avance_path=None, modified_time=""):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ---------- DASHBOARD: KPIs ejecutivos (fila 12) + filtro corte (fila 7) ----------
    db = wb["DASHBOARD"]
    rows = list(db.iter_rows(min_row=1, max_row=30, max_col=12, values_only=True))
    kpis, fecha_corte = {}, ""
    for i, r in enumerate(rows):
        cells = [str(x or "") for x in r]
        if any("VALOR EJECUTADO" in c for c in cells):
            v = rows[i+1]
            kpis = {"valorTotal": f2(v[1]), "ejecutado": f2(v[2]), "progSemana": f2(v[3]),
                    "pctReal": f2(v[4]), "pctProg": f2(v[5]), "spi": f2(v[6])}
        if any("Fecha de Corte" in c for c in cells):
            v = rows[i+1]
            fecha_corte = fecha_str(v[3]) if len(v) > 3 else ""
    kpis["fechaCorte"] = fecha_corte

    # ---------- _DashData: capitulos + curva S semanal ----------
    dd = list(wb["_DashData"].iter_rows(values_only=True))
    capitulos = []
    for r in dd[1:]:
        c0 = str(r[0] or "").strip()
        if c0 and (c0[:1].isdigit() and "." in c0[:3]):
            capitulos.append({"cap": c0, "pctReal": f2(r[1]), "pctProg": f2(r[2])})
    # ---------- Curva S: programado del FLUJO INTER (fila fechas + "% ACUMULADO PROGRAMADO") ----------
    fl = wb["FLUJO DE CAJA_INTER"]
    fl_rows = list(fl.iter_rows(values_only=True))
    fila_fechas = None
    for r in fl_rows[:30]:
        if sum(1 for v in r if isinstance(v, datetime.datetime)) > 20: fila_fechas = r; break
    fila_prog = None
    for r in fl_rows:
        if any("% ACUMULADO PROGRAMADO" in str(c or "") for c in r[:4]): fila_prog = r; break
    semanas, fechas, prog = [], [], []
    fechas_ini = []
    if fila_fechas and fila_prog:
        n = 0
        for j, v in enumerate(fila_fechas):
            if isinstance(v, datetime.datetime):
                n += 1
                semanas.append("S%02d" % n)
                fechas_ini.append(v)
                fin = v + datetime.timedelta(days=6)
                fechas.append(v.strftime("%d-%m-%y") + " a " + fin.strftime("%d-%m-%y"))
                prog.append(f2(fila_prog[j]) if j < len(fila_prog) else 0.0)

    # ---------- Curva S real: AVANCE DIARIO (cantidades diarias x vrUnit) ----------
    real = [0.0] * len(semanas)
    if avance_path:
        wa = openpyxl.load_workbook(avance_path, read_only=True, data_only=True)
        ws = wa["AVANCE DIARIO"]
        arows = list(ws.iter_rows(values_only=True))
        hdr_i = None
        for i, row in enumerate(arows[:12]):
            if sum(1 for v in row if isinstance(v, datetime.datetime)) > 50: hdr_i = i; break
        if hdr_i is not None:
            hdr = arows[hdr_i]
            fcols = [(j, v) for j, v in enumerate(hdr) if isinstance(v, datetime.datetime)]
            # vrUnit por fila = col 4 (0-based); cantidades en cols de fecha
            por_dia = {}
            for row in arows[hdr_i + 2:]:
                vu = f2(row[4]) if len(row) > 4 else 0.0
                if vu <= 0: continue
                for j, f in fcols:
                    c = row[j] if j < len(row) else None
                    if isinstance(c, (int, float)) and c:
                        por_dia[f.date()] = por_dia.get(f.date(), 0.0) + c * vu
            total = kpis.get("valorTotal", 0) or 1
            # bucket semanal alineado a fechas_ini (jueves a miercoles)
            acum = 0.0
            dias_orden = sorted(por_dia)
            di = 0
            for k in range(len(fechas_ini)):
                fin_sem = (fechas_ini[k] + datetime.timedelta(days=6)).date()
                while di < len(dias_orden) and dias_orden[di] <= fin_sem:
                    acum += por_dia[dias_orden[di]]; di += 1
                real[k] = acum / total
            # Escalar al total oficial del DASHBOARD (las hojas usan precios distintos;
            # la FORMA semanal viene del AVANCE DIARIO, el NIVEL lo fija el DASHBOARD)
            pct_oficial = kpis.get("pctReal", 0)
            ult_val = real[-1] if real and real[-1] else max((v for v in real if v), default=0)
            if pct_oficial > 0 and ult_val > 0:
                k_esc = pct_oficial / ult_val
                real = [ (v * k_esc if v is not None else v) for v in real ]
            # despues de la ultima semana con datos -> null
            if dias_orden:
                ult = dias_orden[-1]
                for k in range(len(fechas_ini)):
                    if fechas_ini[k].date() > ult: real[k] = None
    else:
        real = [None] * len(semanas)

    # ---------- CRONORGRAMA_AVANCE: 98 actividades ----------
    cr = wb["CRONORGRAMA_AVANCE"]
    acts, hdr_found, idx = [], False, 0
    for r in cr.iter_rows(min_row=12, max_col=20, values_only=True):
        if not hdr_found:
            if r[0] and "Cap" in str(r[0]): hdr_found = True
            continue
        if not r[1] or not str(r[1]).strip(): continue
        item = str(r[1]).strip()
        if not item[0].isdigit(): continue
        idx += 1
        acts.append({
            "idx": idx, "cap": str(r[0] or "").strip(), "item": item,
            "desc": str(r[2] or "").strip(), "resp": str(r[3] or "").strip(),
            "cantEjec": f2(r[4]), "pctAcum": f2(r[7]), "unidad": str(r[8] or "").strip(),
            "cantContr": f2(r[9]), "cuadrillas": f2(r[10]), "tDias": f2(r[14]),
            "vrUnit": f2(r[15]), "vrEjec": f2(r[16]), "vrTotal": f2(r[17]),
            "fIni": fecha_str(r[18]), "fFin": fecha_str(r[19]),
        })

    out = {
        "fuente": "github-actions",
        "_generado": datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "modifiedTimeDrive": modified_time,
        "proyecto": "REPOSICION DE REDES DE ALCANTARILLADO - COMUNEROS",
        "ejecutadoGlobal": kpis.get("ejecutado", 0),
        "dashboardKpis": kpis,
        "capitulos": capitulos,
        "actividades": acts,
        "curvaS": {"semanas": semanas, "fechas": fechas, "programado": prog, "real": real},
    }
    return out

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else None
    avance = sys.argv[2] if len(sys.argv) > 2 else None
    meta = {}
    if not xlsx:
        xlsx = "/tmp/analisis_dl.xlsx"; avance = "/tmp/avance_dl.xlsx"
        meta = descargar_drive(ANALISIS_ID, xlsx)
        m2 = descargar_drive(AVANCE_ID, avance)
        print("Descargados:", meta.get("modifiedTime"), m2.get("modifiedTime"))
    data = construir(xlsx, avance, meta.get("modifiedTime", ""))
    prev = None
    try: prev = json.load(open(OUT, encoding="utf-8"))
    except Exception: pass
    if prev and prev.get("ejecutadoGlobal") == data["ejecutadoGlobal"] and \
       prev.get("curvaS", {}).get("real") == data["curvaS"]["real"] and \
       len(prev.get("actividades", [])) == len(data["actividades"]) and \
       prev.get("dashboardKpis") == data["dashboardKpis"]:
        print("Sin cambios de datos; no se reescribe.")
        return 0
    json.dump(data, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
    print("obra_data.json: %d actividades, %d semanas, ejecutado=%.0f (%.2f%%)" %
          (len(data["actividades"]), len(data["curvaS"]["semanas"]),
           data["ejecutadoGlobal"], data["dashboardKpis"].get("pctReal", 0) * 100))
    return 0

if __name__ == "__main__":
    sys.exit(main())
